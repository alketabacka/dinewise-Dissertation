from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import pandas as pd
import openai
import os
import time
import logging
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openai.error import RateLimitError, OpenAIError
import bcrypt
from dotenv import load_dotenv

logging.basicConfig(level=logging.DEBUG)
app = Flask(__name__)
app.secret_key = os.urandom(24)

load_dotenv()
# Initialize OpenAI API
openai.api_key = os.getenv('OPENAI_API_KEY')

# Load restaurant information from Excel
restaurant_data = pd.read_excel('Restaurant information.xlsx')

@app.route('/')
def index():
    logging.debug("Rendering index page")
    return render_template('index.html')

@app.route('/register')
def register():
    logging.debug("Rendering register page")
    return render_template('register.html')

@app.route('/api/register', methods=['POST'])
def api_register():
    try:
        data = request.json
        firstname = data.get('firstname')
        lastname = data.get('lastname')
        email = data.get('email')
        password = data.get('password')
        
        # Hash the password
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Path to the Excel file
        excel_path = 'Book1.xlsx'
        
        # Read the existing Excel file
        try:
            df = pd.read_excel(excel_path)
        except FileNotFoundError:
            df = pd.DataFrame(columns=['First Name', 'Last Name', 'Email', 'Password'])

        # Check if the email already exists
        if email in df['Email'].values:
            return jsonify({'status': 'error', 'message': 'This email is already registered. Please use a different email.'})

        # Create a new DataFrame with the same columns
        new_data = pd.DataFrame([[firstname, lastname, email, hashed_password]], 
                                columns=['First Name', 'Last Name', 'Email', 'Password'])
        
        # Append the new data
        df = pd.concat([df, new_data], ignore_index=True)

        # Write the data back to the Excel file
        df.to_excel(excel_path, index=False)
        
        # Automatically log in the user after registration
        session['user'] = {'First Name': firstname, 'Last Name': lastname, 'Email': email}

        return jsonify({'status': 'success', 'message': 'Registration successful.'})
    except Exception as e:
        logging.error(f"An error occurred during registration: {e}")
        return jsonify({'status': 'error', 'message': 'An unexpected error occurred. Please try again later.'})

@app.route('/login', methods=['GET'])
def login():
    logging.debug("Rendering login page")
    return render_template('login.html')

@app.route('/admin')
def admin():
    logging.debug("Rendering admin page")
    return render_template('admin.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')
        
        # Path to the Excel file
        excel_path = 'Book1.xlsx'
        
        # Read the existing Excel file
        try:
            df = pd.read_excel(excel_path)
        except FileNotFoundError:
            return jsonify({'status': 'error', 'message': 'No registered users found.'})
        
        # Check if the user exists and the password matches
        user = df[df['Email'] == email]
        if not user.empty:
            stored_hashed_password = user.iloc[0]['Password']
            if bcrypt.checkpw(password.encode('utf-8'), stored_hashed_password.encode('utf-8')):
                session['user'] = user.iloc[0].to_dict()
                logging.debug("User logged in successfully")
                return jsonify({'status': 'success', 'message': 'Login successful.'})
            else:
                return jsonify({'status': 'error', 'message': 'Invalid credentials. Please try again.'})
        else:
            return jsonify({'status': 'error', 'message': 'Invalid credentials. Please try again.'})
    except Exception as e:
        logging.error(f"An error occurred during login: {e}")
        return jsonify({'status': 'error', 'message': 'An unexpected error occurred. Please try again later.'})

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('index'))

@app.route('/chatbot', methods=['GET', 'POST'])
def chatbot():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        user_message = request.form['message']
        
        response = handle_user_message(user_message)
        
        return render_template('chatbot.html', user_message=user_message, bot_response=response)
    
    return render_template('chatbot.html')

def handle_user_message(user_message):
    response = ""
    if 'restaurants in' in user_message.lower():
        cuisine, city = extract_cuisine_and_city(user_message)
        response = get_restaurants_response(cuisine, city)
    elif 'menu for' in user_message.lower():
        restaurant_name = user_message.split('menu for ')[1].strip()
        response = get_menu_response(restaurant_name)
    elif 'reviews for' in user_message.lower():
        restaurant_name = user_message.split('reviews for ')[1].strip()
        response = get_reviews_response(restaurant_name)
    elif 'location of' in user_message.lower() or 'postcode of' in user_message.lower():
        restaurant_name = user_message.split('location of ')[1].strip() if 'location of' in user_message.lower() else user_message.split('postcode of ')[1].strip()
        response = get_location_response(restaurant_name)
    elif 'near' in user_message.lower():
        area = user_message.split('near ')[1].strip()
        response = get_nearby_restaurants_response(area)
    else:
        response = get_openai_response(user_message)
    
    return response

def extract_cuisine_and_city(message):
    parts = message.lower().split('restaurants in ')
    if len(parts) > 1:
        city = parts[1].strip()
        cuisine = parts[0].replace('restaurants', '').strip()
        return cuisine, city
    return None, None

def get_restaurants_response(cuisine, city):
    if cuisine and city:
        filtered_restaurants = restaurant_data[
            (restaurant_data['Location'].str.contains(city, case=False, na=False)) &
            (restaurant_data['Type of restaurant'].str.contains(cuisine, case=False, na=False))
        ]
        if not filtered_restaurants.empty:
            unique_restaurants = filtered_restaurants['Restaurant Name'].unique()
            return f"Here are some {cuisine.capitalize()} restaurants in {city}:\n" + ", ".join(unique_restaurants)
        else:
            return f"Sorry, I couldn't find any {cuisine} restaurants in {city}. Perhaps try another city or cuisine?"
    elif city:
        restaurants = restaurant_data[restaurant_data['Location'].str.contains(city, case=False, na=False)]
        if not restaurants.empty:
            unique_restaurants = restaurants['Restaurant Name'].unique()
            return f"Here are some restaurants in {city}:\n" + ", ".join(unique_restaurants)
        else:
            return f"Sorry, I couldn't find any restaurants in {city}. Perhaps try another city?"
    else:
        return "Please specify a city to search for restaurants."

def get_nearby_restaurants_response(area):
    if area:
        nearby_restaurants = restaurant_data[restaurant_data['Location'].str.contains(area, case=False, na=False)]
        if not nearby_restaurants.empty():
            unique_restaurants = nearby_restaurants['Restaurant Name'].unique()
            return f"Here are some restaurants near {area}:\n" + ", ".join(unique_restaurants)
        else:
            return f"Sorry, I couldn't find any restaurants near {area}. Perhaps try another area?"
    else:
        return "Please specify an area to search for nearby restaurants."

def get_menu_response(restaurant_name):
    if restaurant_name:
        menus = restaurant_data[restaurant_data['Restaurant Name'].str.contains(restaurant_name, case=False, na=False)]
        if not menus.empty:
            # Copy to avoid SettingWithCopyWarning
            menus_copy = menus.copy()
            # Convert all Menu and Price values to strings, replacing NaNs with empty strings
            menus_copy['Menu'] = menus_copy['Menu'].fillna('').astype(str)
            menus_copy['Price'] = menus_copy['Price'].fillna(0).astype(float)
            menu_list = [f"{row['Menu']} - £{row['Price']:.2f}" for _, row in menus_copy.iterrows()]
            combined_menu = "\n".join(menu_list)
            return f"The menu for {restaurant_name} is:\n{combined_menu}"
        else:
            return f"Sorry, I couldn't find the menu for {restaurant_name}. Perhaps try another restaurant?"
    else:
        return "Please specify a restaurant to get the menu."


def get_reviews_response(restaurant_name):
    if restaurant_name:
        reviews = restaurant_data[restaurant_data['Restaurant Name'].str.contains(restaurant_name, case=False, na=False)]
        if not reviews.empty():
            first_review = reviews['Overall reviews'].iloc[0]
            return f"The review for {restaurant_name} is:\n{first_review}"
        else:
            return f"Sorry, I couldn't find any reviews for {restaurant_name}. Perhaps try another restaurant?"
    else:
        return "Please specify a restaurant to get the reviews."

def get_location_response(restaurant_name):
    if restaurant_name:
        location_data = restaurant_data[restaurant_data['Restaurant Name'].str.contains(restaurant_name, case=False, na=False)]
        if not location_data.empty():
            location = location_data['Location'].iloc[0]
            postcode = location_data['Post-Code'].iloc[0]
            return f"The location of {restaurant_name} is {location} and the postcode is {postcode}."
        else:
            return f"Sorry, I couldn't find the location or postcode for {restaurant_name}. Perhaps try another restaurant?"
    else:
        return "Please specify a restaurant to get the location or postcode."


def get_openai_response(user_message):
    max_retries = 3
    retry_delay = 1  # Initial delay in seconds

    for attempt in range(max_retries):
        try:
            logging.info(f"Attempt {attempt + 1} to call OpenAI API")
            openai_response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": user_message}],
                max_tokens=150
            )
            logging.info(f"OpenAI response: {openai_response}")
            return openai_response.choices[0].message['content'].strip()
        except RateLimitError as e:
            logging.error(f"OpenAI rate limit exceeded: {e}")
            return "I'm currently unavailable due to high demand. Please try again later."
        except OpenAIError as e:
            logging.error(f"OpenAI API call failed: {e}")
            time.sleep(retry_delay)
            retry_delay *= 2  # Exponential backoff for retries
        except Exception as e:
            logging.error(f"An unexpected error occurred: {e}")
            return "An unexpected error occurred. Please try again later."
    return "I'm currently experiencing issues. Please try again later."

@app.route('/booking')
def booking():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user = session['user']
    fullname = f"{user['First Name']} {user['Last Name']}"
    email = user['Email']
    phone = user.get('Phone Number', '')  # Adjust this based on your user data structure
    
    # Fetch unique restaurant names from your data
    restaurant_names = restaurant_data['Restaurant Name'].unique().tolist()
    
    return render_template('booking.html', fullname=fullname, email=email, phone=phone, restaurant_names=restaurant_names)

@app.route('/get_menu', methods=['POST'])
def get_menu():
    restaurant_name = request.form['restaurant']
    menus = restaurant_data[restaurant_data['Restaurant Name'].str.contains(restaurant_name, case=False, na=False)]
    
    if not menus.empty:
        # Copy to avoid SettingWithCopyWarning
        menus_copy = menus.copy()
        # Convert all Menu and Price values to strings, replacing NaNs with empty strings
        menus_copy['Menu'] = menus_copy['Menu'].fillna('').astype(str)
        menus_copy['Price'] = menus_copy['Price'].fillna(0).astype(float)
        menu_list = [f"{row['Menu']} - £{row['Price']:.2f}" for _, row in menus_copy.iterrows()]
        return jsonify(menu_list)
    else:
        return jsonify([])  # Return an empty list if no menu is found

@app.route('/book_table', methods=['POST'])
def book_table():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user = session['user']
    name = request.form.get('name')
    restaurant = request.form.get('restaurant')
    date = request.form.get('date')
    time = request.form.get('time')
    guests = int(request.form.get('guests'))
    desires = request.form.get('desires')
    email = request.form.get('email')
    phone = request.form.get('phone')
    
    # Path to the Excel file
    excel_path = 'the reservation book.xlsx'
    
    # Load the existing Excel file
    book = openpyxl.load_workbook(excel_path)
    sheet = book.active

    # Check if there are already 3 bookings for the same restaurant at the same time
    existing_bookings = 0
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
        if row[2] == restaurant and row[4] == date and row[5] == time:
            existing_bookings += 1

    if existing_bookings >= 3:
        return jsonify({"status": "error", "message": "Sorry, the maximum number of bookings for this time slot has been reached."})

    # Ensure the new "Person Number" column exists
    if 'Person Number' not in [cell.value for cell in sheet[1]]:
        sheet.insert_cols(9)
        sheet.cell(row=1, column=9, value="Person Number")

    total_price_all_guests = 0  # Initialize the total price for all guests

    # Save details for each guest, including the user
    for i in range(1, guests + 1):
        selected_items = request.form.get(f'selectedItemsPerson{i}')
        
        # Convert the JSON string back to a list
        order_list = eval(selected_items) if selected_items else []
        order = ', '.join([item['name'] for item in order_list])
        total_price = sum(item['price'] for item in order_list)
        special_desires = request.form.get(f'desiresPerson{i}')
        
        total_price_all_guests += total_price  # Accumulate the total price

        # Find the next empty row
        next_row = sheet.max_row + 1
        # Fill the row with form data
        sheet.cell(row=next_row, column=4, value=name)
        sheet.cell(row=next_row, column=3, value=restaurant)
        sheet.cell(row=next_row, column=5, value=date)
        sheet.cell(row=next_row, column=6, value=time)
        sheet.cell(row=next_row, column=11, value=guests)
        sheet.cell(row=next_row, column=8, value=special_desires)
        sheet.cell(row=next_row, column=12, value=email)
        sheet.cell(row=next_row, column=10, value=phone)
        sheet.cell(row=next_row, column=9, value=f"Person {i}")
        sheet.cell(row=next_row, column=7, value=f"{order}")
        sheet.cell(row=next_row, column=13, value=total_price)

    # Save the Excel file
    book.save(excel_path)
    
    # Send confirmation email
    send_confirmation_email(email, name, restaurant, date, time, guests, total_price_all_guests)

    return jsonify({"status": "success"})

@app.route('/manage_bookings')
def manage_bookings():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user_email = session['user']['Email']
    
    # Path to the Excel file
    excel_path = 'the reservation book.xlsx'
    
    # Load the existing Excel file
    book = openpyxl.load_workbook(excel_path)
    sheet = book.active
    
    # Find bookings for the logged-in user
    user_bookings = []
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
        if row[11] == user_email:  # Assuming email is in the 12th column (index 11)
            user_bookings.append({
                'name': row[3],
                'restaurant': row[2],
                'date': row[4],
                'time': row[5],
                'guests': row[9],
                'desires': row[7],
                'email': row[11],
                'phone': row[8],
                'order': row[6],
                'total_price': row[12]
            })
    
    print("User bookings:", user_bookings)  # Debugging line
    
    return render_template('manage_bookings.html', bookings=user_bookings)
@app.route('/cancel_booking', methods=['POST'])
def cancel_booking():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user_email = session['user']['Email']
    booking_to_cancel = request.form.get('booking_to_cancel')
    
    # Path to the Excel file
    excel_path = 'the reservation book.xlsx'
    
    # Load the existing Excel file
    try:
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
    except Exception as e:
        return f"Error loading Excel file: {str(e)}", 500
    
    # Find and delete the booking
    booking_found = False
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
        if row[10].value == user_email and row[3].value == booking_to_cancel:
            sheet.delete_rows(row[0].row)
            booking_found = True
            break
    
    if booking_found:
        try:
            book.save(excel_path)
        except Exception as e:
            return f"Error saving Excel file: {str(e)}", 500
        
        return redirect(url_for('manage_bookings'))
    
    return 'Booking not found', 404


def send_confirmation_email(email, name, restaurant, date, time, guests, total_price_all_guests):
    sender_email = "futurestudiesal@gmail.com"
    sender_password = "smouaqyorbjrsicd"
    subject = "Booking Confirmation"
    body = f"""
    Dear {name},

    Your booking at {restaurant} has been confirmed.

    Details:
    Date: {date}
    Time: {time}
    Guests: {guests}
    Total Price: £{total_price_all_guests:.2f}

    Individual Orders:
    """
    
    for i in range(1, guests + 1):
        selected_items_person = eval(request.form.get(f'selectedItemsPerson{i}')) if request.form.get(f'selectedItemsPerson{i}') else []
        order = ', '.join([item['name'] for item in selected_items_person])
        special_desires = request.form.get(f'desiresPerson{i}')
        body += f"""
        Person {i}:
        Order: {order}
        Special Desires: {special_desires}
        """
    
    body += """
    Thank you for your booking.

    Best regards,
    Your {restaurant} Team
    """
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    
    try:
        logging.info("Connecting to the SMTP server...")
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        logging.info("Logging in to the SMTP server...")
        server.login(sender_email, sender_password)
        logging.info("Sending the email...")
        text = msg.as_string()
        server.sendmail(sender_email, email, text)
        server.quit()
        logging.info("Email sent successfully")
    except smtplib.SMTPAuthenticationError:
        logging.error("Failed to authenticate with the SMTP server. Check your username and password.")
    except smtplib.SMTPConnectError:
        logging.error("Failed to connect to the SMTP server. Check your network connection.")
    except smtplib.SMTPException as e:
        logging.error(f"An SMTP error occurred: {e}")
    except Exception as e:
        logging.error(f"An error occurred while sending the email: {e}")

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    error_message = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Static admin credentials
        admin_username = "admin"
        admin_password = "admin123"

        if username == admin_username and password == admin_password:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            error_message = 'Invalid credentials. Please try again.'

    return render_template('admin.html', error_message=error_message)

@app.route('/view_register_book')
def view_register_book():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    excel_path = 'Book1.xlsx'
    df = pd.read_excel(excel_path)
    register_book = df.to_dict(orient='records')

    return render_template('view_register_book.html', register_book=register_book)

@app.route('/view_restaurant_information', methods=['GET', 'POST'])
def view_restaurant_information():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    excel_path = 'Restaurant information.xlsx'
    df = pd.read_excel(excel_path)

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            new_restaurant = {
                'Restaurant Name': request.form['name'],
                'Type of restaurant': request.form['type'],
                'Location': request.form['location'],
                'Post-Code': request.form['postcode'],
                'Overall reviews': request.form['reviews'],
                'Menu': request.form['menu'],
                'Price': request.form['price']
            }
            df = pd.concat([df, pd.DataFrame([new_restaurant])], ignore_index=True)
            df.to_excel(excel_path, index=False)
        
        elif action == 'delete':
            restaurant_name = request.form['name']
            df = df[df['Restaurant Name'] != restaurant_name]
            df.to_excel(excel_path, index=False)

        elif action == 'update':
            restaurant_name = request.form['name']
            for index, row in df.iterrows():
                if row['Restaurant Name'] == restaurant_name:
                    df.at[index, 'Type of restaurant'] = request.form['type']
                    df.at[index, 'Location'] = request.form['location']
                    df.at[index, 'Post-Code'] = request.form['postcode']
                    df.at[index, 'Overall reviews'] = request.form['reviews']
                    df.at[index, 'Menu'] = request.form['menu']
                    df.at[index, 'Price'] = request.form['price']
                    df.to_excel(excel_path, index=False)
                    break

    restaurant_information = df.to_dict(orient='records')
    return render_template('view_restaurant_information.html', restaurant_information=restaurant_information)

@app.route('/view_reservations')
def view_reservations():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    # Path to the Excel file
    excel_path = 'the reservation book.xlsx'

    # Load the existing Excel file
    df = pd.read_excel(excel_path)

    # Convert DataFrame to a list of dictionaries
    reservations = df.to_dict(orient='records')

    return render_template('view_reservations.html', reservations=reservations)

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'admin' in session:
        return render_template('admin_dashboard.html')
    else:
        return redirect(url_for('admin_login'))

@app.route('/logout_admin')
def logout_admin():
    session.pop('admin', None)
    return redirect(url_for('admin_login'))

if __name__ == '__main__':
    app.run(debug=True)
